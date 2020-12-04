import openpyxl
from datetime import datetime


now_date = datetime.now()
now_day = now_date.day
now_month = now_date.month
now_year = now_date.year
now_minute = now_date.minute


class ToExcel:
    """Запись в файд Excel"""
    def __init__(self, file_name):
        self.__now_date = f"{now_day}.{now_month}.{now_year}M{now_minute}"
        self.__file_name = f"result/{file_name}.xlsx"
        f = open(self.__file_name, "wb")
        f.close()
        self.__wb = openpyxl.Workbook()
        self.__wb.create_sheet(title=self.__now_date, index=0)
        self.__sheet = self.__wb[self.__now_date]

    # вставка заголовков столбцов со списка
    def write_to_excel_name_col_list(self, lst):
        for i in range(len(lst)):
            cell = self.__sheet.cell(row=1, column=i+1)
            cell.value = lst[i]

    # вставка в ячейки все значения ключей из словаря
    def write_to_excel_value_from_dict(self, dct, row, col=1):
        for key in dct:
            cell = self.__sheet.cell(row=row, column=col)
            cell.value = dct.get(key)
            col += 1

    # сохранение файла
    def save_excel(self):
        self.__wb.save(self.__file_name)

    # вставка заголовков столбцов со словаря
    def write_to_excel_name_col_dict(self, dct, col=1):
        for key in dct:
            cell = self.__sheet.cell(row=1, column=col)
            cell.value = key
            col += 1

    # вставка соответствующего значения из словаря в нужный столбец
    def write_to_excel_value(self, lst, dct, row):
        for key in lst:
            if key.lower() in dct:
                cell = self.__sheet.cell(row=row, column=lst.index(key) + 1)
                cell.value = dct.get(key.lower())


class FromExcel:
    """Чтение с файла Excel"""
    def __init__(self, filename):
        self.__wb = openpyxl.load_workbook(filename)
        self.__sheet = self.__wb["form"]

    # чтение данных по столбцам
    def __read_from_excel(self, col, line_start, line_stop=None):
        self.__result = list()
        self.__line = line_start
        while True:
            value = self.__sheet[f"{col}{self.__line}"].value
            if value is None or line_stop == self.__line:
                break
            self.__result.append(value)
            self.__line += 1
        return self.__result

    # чтение ссылки
    def read_cell_str(self, col, line):
        return self.__read_from_excel(col, line, line+1)[0]

    # чтение пунктов
    def read_col_in_list(self, col, line_start):
        return self.__read_from_excel(col, line_start)
